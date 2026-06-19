import streamlit as st
import pandas as pd
from playwright.sync_api import sync_playwright
from st_aggrid import AgGrid, GridOptionsBuilder, DataReturnMode
import openpyxl
import os
import time
import re
from datetime import datetime

# --- Setup ---
st.set_page_config(layout="wide", page_title="Cortex Guide Downloader")

# Header with Refresh Logic
col_title, col_ref = st.columns([0.8, 0.2])
col_title.title("📚 Cortex Admin Guide Manager")

# --- Persistent State Management ---
if "logs" not in st.session_state:
    st.session_state.logs = []
if "selected_excel_rows" not in st.session_state:
    st.session_state.selected_excel_rows = []
if "grid_key" not in st.session_state:
    st.session_state.grid_key = 0
if "custom_filenames" not in st.session_state:
    st.session_state.custom_filenames = {}

# --- REFRESH ACTION ---
if col_ref.button("🔄 Reload from Excel"):
    st.session_state.selected_excel_rows = []
    st.session_state.custom_filenames = {}
    st.session_state.grid_key += 1
    print("\n--- MANUAL RELOAD TRIGGERED ---")
    st.rerun()

EXCEL_FILE = "sources.xlsx"

def add_log(message):
    """Prints to BOTH the Streamlit UI and the Terminal Window"""
    timestamp = datetime.now().strftime("%H:%M:%S")
    formatted_msg = f"[{timestamp}] {message}"
    
    # 1. Update Streamlit Session State
    st.session_state.logs.append(formatted_msg)
    st.session_state.logs = st.session_state.logs[-100:]
    
    # 2. Print to Terminal (Standard Output)
    print(formatted_msg)

def load_excel_data():
    if not os.path.exists(EXCEL_FILE):
        st.error(f"Could not find {EXCEL_FILE}.")
        return pd.DataFrame()
    
    print(f"Reading file: {EXCEL_FILE}...")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1] if cell.value is not None]
    if not headers or 'File Title' not in headers:
        return pd.DataFrame()
        
    title_idx = headers.index('File Title')
    data = []
    for i, row in enumerate(ws.iter_rows(min_row=2)):
        if not any(cell.value for cell in row): continue
        row_dict = {headers[j]: row[j].value for j in range(len(headers))}
        title_cell = row[title_idx]
        
        extracted_url = ""
        if title_cell.hyperlink: 
            extracted_url = title_cell.hyperlink.target
        elif isinstance(title_cell.value, str) and title_cell.value.strip().lower().startswith("http"):
            extracted_url = title_cell.value.strip()
            
        row_dict['Extracted URL'] = extracted_url
        row_dict['Excel Row Index'] = i + 2 
        if not row_dict.get('Status'): row_dict['Status'] = 'Pending'
        
        base_title = str(row_dict.get('File Title', 'Untitled')).split('.pdf')[0].split('.PDF')[0]
        clean_title = re.sub(r'[\\/*?:"<>|]', "", base_title)
        row_dict['Download Filename ✏️'] = st.session_state.custom_filenames.get(row_dict['Excel Row Index'], clean_title)
        data.append(row_dict)
    return pd.DataFrame(data)

def update_excel_cell(row_index, col_name, new_value):
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        if col_name in headers:
            ws.cell(row=row_index, column=headers.index(col_name) + 1).value = new_value
            wb.save(EXCEL_FILE)
            print(f"Excel row {row_index} updated: {col_name} -> {new_value}")
    except Exception as e:
        print(f"FAILED TO UPDATE EXCEL: {e}")

def format_date(raw_date_string):
    clean_string = re.sub(r'[^\w\s,-]', '', raw_date_string).strip()
    try: 
        return pd.to_datetime(clean_string).strftime("%b %d, %Y")
    except:
        try: 
            return pd.to_datetime(clean_string, dayfirst=True).strftime("%b %d, %Y")
        except: 
            return raw_date_string

df = load_excel_data()

if not df.empty:
    st.markdown("### 1. Select Guides to Process")
    search_text = st.text_input("🔍 Quick Search:", "")
    
    excel_to_zero_idx = {row['Excel Row Index']: i for i, row in enumerate(df.to_dict('records'))}
    pre_selected_indices = [excel_to_zero_idx[er] for er in st.session_state.selected_excel_rows if er in excel_to_zero_idx]
    
    gb = GridOptionsBuilder.from_dataframe(df)
    gb.configure_selection('multiple', use_checkbox=True, header_checkbox=True, pre_selected_rows=pre_selected_indices)
    gb.configure_column("Extracted URL", hide=True)
    gb.configure_column("Excel Row Index", hide=True)
    gb.configure_default_column(editable=False)
    gb.configure_column("Download Filename ✏️", editable=True, flex=1)
    gb.configure_grid_options(quickFilterText=search_text)
    
    grid_response = AgGrid(
        df, 
        gridOptions=gb.build(), 
        data_return_mode=DataReturnMode.AS_INPUT, 
        update_on=['selectionChanged', 'cellValueChanged'], 
        theme='alpine', 
        height=400, 
        width='100%', 
        key=f'grid_{st.session_state.grid_key}'
    )
    
    if isinstance(grid_response['selected_rows'], pd.DataFrame):
        selected_records = grid_response['selected_rows'].to_dict('records')
    else:
        selected_records = grid_response['selected_rows'] or []
        
    st.session_state.selected_excel_rows = [r['Excel Row Index'] for r in selected_records if 'Excel Row Index' in r]
    
    if isinstance(grid_response['data'], pd.DataFrame):
        full_grid_data = grid_response['data'].to_dict('records')
        for r in full_grid_data: 
            st.session_state.custom_filenames[r['Excel Row Index']] = r['Download Filename ✏️']
            
    items = [row for row in full_grid_data if row['Excel Row Index'] in st.session_state.selected_excel_rows]

    st.markdown("---")
    st.markdown("### 2. Run Automations")
    col1, col2 = st.columns(2)
    log_container = st.empty()
    log_container.code('\n'.join(st.session_state.logs) if st.session_state.logs else "System Idle...", language='bash')

    # --- DATE FETCH ---
    if col1.button("🔍 Fetch Dates"):
        if not items: st.warning("Select rows first!")
        else:
            add_log("=== STARTING DATE FETCH JOB (HEADLESS) ===")
            with sync_playwright() as p:
                browser = p.chromium.launch(headless=True)
                page = browser.new_page()
                for row in items:
                    add_log(f"Fetching Date: {row['File Title']}")
                    log_container.code('\n'.join(st.session_state.logs), language='bash')
                    try:
                        if not row['Extracted URL']: 
                            add_log("No URL found, skipping...")
                            continue
                        page.goto(row['Extracted URL'], wait_until="domcontentloaded", timeout=90000)
                        time.sleep(6) 
                        shadow_js = "() => { let t=[]; function w(n){ if(n.shadowRoot)w(n.shadowRoot); if(n.nodeType===3){let v=n.nodeValue.trim(); if(v)t.push(v);} for(let c of n.childNodes)w(c); } w(document.body); return t.join(' '); }"
                        body_text = page.evaluate(shadow_js)
                        match = re.search(r'(?:Publication Date:|Last publication:|Last date published|ft:lastEdition:|ft:lastPublication:|Updated:|Published:)[\s]*([0-9]{2}-[0-9]{2}-[0-9]{4}|[0-9]{4}-[0-9]{2}-[0-9]{2}|[A-Za-z]{3,9}\s+[0-9]{1,2},?\s+[0-9]{4})', body_text, re.IGNORECASE)
                        if match:
                            d = format_date(match.group(1).strip())
                            update_excel_cell(row['Excel Row Index'], "Version", d)
                            update_excel_cell(row['Excel Row Index'], "Status", "✅ Date Updated")
                            add_log(f"✅ Success: {d}")
                        else: add_log(f"⚠️ Regex miss for '{row['File Title']}'")
                        log_container.code('\n'.join(st.session_state.logs), language='bash')
                    except Exception as e:
                        add_log(f"ERROR: {str(e)[:60]}")
                browser.close()
            add_log("=== DATE FETCH JOB COMPLETE ===")
            st.rerun()

    # --- PDF DOWNLOAD ---
    if col2.button("⬇️ Download PDFs"):
        if not items: st.warning("Select rows first!")
        else:
            add_log("=== STARTING HEAVYWEIGHT DOWNLOAD JOB ===")
            with sync_playwright() as p:
                browser = p.chromium.launch(headless=True)
                context = browser.new_context()
                context.on("page", lambda new_page: new_page.add_init_script("window.print = () => {}; window.close = () => {};"))
                page = context.new_page()
                
                for row in items:
                    if row["Status"] == "✅ Downloaded":
                        add_log(f"⏭️ SKIPPING: '{row['File Title']}' is already marked as Downloaded.")
                        log_container.code('\n'.join(st.session_state.logs), language='bash')
                        continue
                    try:
                        add_log(f"🚀 Processing: {row['Download Filename ✏️']}")
                        log_container.code('\n'.join(st.session_state.logs), language='bash')
                        
                        page.goto(row['Extracted URL'], wait_until="networkidle", timeout=120000)
                        time.sleep(8)
                        
                        # STEP 1: Click Print
                        add_log("Step 1: Piercing DOM to click Print Map icon...")
                        page.evaluate("() => { function w(n){ if(n.shadowRoot)w(n.shadowRoot); if(n.tagName==='FT-READER-ACTION'&&n.getAttribute('action')==='print-map'){let b=n.shadowRoot?n.shadowRoot.querySelector('button'):null; if(b)b.click(); else n.click();} for(let c of n.childNodes)w(c); } w(document.body); }")
                        time.sleep(6)
                        
                        # STEP 2: Select All
                        add_log("Step 2: Locating Modal and clicking 'Select All'...")
                        page.evaluate("() => { function w(n){ if(n.shadowRoot)w(n.shadowRoot); if(n.tagName==='FT-CHECKBOX'){let i=n.shadowRoot?n.shadowRoot.querySelector('input'):null; if(i&&!i.checked)i.click();} for(let c of n.childNodes)w(c); } w(document.body); }")
                        time.sleep(5)
                        
                        # STEP 3: Confirm & Catch Tab
                        target_page = page
                        try:
                            add_log("Step 3: Triggering PDF compilation...")
                            with context.expect_page(timeout=300000) as new_page_info:
                                page.evaluate("() => { let btns=[]; function w(n){ if(n.shadowRoot)w(n.shadowRoot); if(n.tagName==='FT-BUTTON'||n.tagName==='BUTTON'){let t=n.textContent.trim().toLowerCase(); if(t==='print'||t==='print topics')btns.push(n);} for(let c of n.childNodes)w(c); } w(document.body); let t=btns[btns.length-1]; if(t){let b=t.shadowRoot?t.shadowRoot.querySelector('button'):null; if(b)b.click(); else t.click();} }")
                            
                            target_page = new_page_info.value
                            add_log("CRITICAL: Compiling massive document. Browser is organizational background now (Waiting up to 5 mins)...")
                            log_container.code('\n'.join(st.session_state.logs), language='bash')
                            
                            # HEARTBEAT LOGS
                            target_page.wait_for_load_state("networkidle", timeout=300000)
                            add_log("Compilation phase 1 complete (Network idle reached).")
                            time.sleep(15) 
                            add_log("Compilation phase 2 complete (Stability buffer finished).")
                        except Exception as e:
                            add_log(f"Fallback mode triggered: Ghost tab not detected ({str(e)[:30]})")
                            time.sleep(20)

                        # SAFE WIDGET ASSASSIN
                        add_log("Cleaning PDF: Removing floating widgets...")
                        try:
                            if not target_page.is_closed():
                                target_page.evaluate('''() => {
                                    const style = document.createElement('style');
                                    style.innerHTML = `ft-feedback-button, ft-scroll-to-top, .ft-floating-button, 
                                                       [class*="feedback"], [id*="feedback"], [class*="widget"], #pendo-base { display: none !important; }`;
                                    document.head.appendChild(style);
                                    function hideFixed(node) {
                                        if (node.shadowRoot) { try { node.shadowRoot.appendChild(style.cloneNode(true)); } catch(e){} hideFixed(node.shadowRoot); }
                                        if (node.nodeType === 1) { 
                                            const comp = window.getComputedStyle(node);
                                            if (comp.position === 'fixed' || comp.position === 'sticky') node.style.display = 'none';
                                        }
                                        for (let child of node.childNodes) hideFixed(child);
                                    }
                                    hideFixed(document.body);
                                }''')
                                time.sleep(5)
                        except Exception as e: 
                            print(f"CSS Injection Warning: {e}")
                        
                        add_log(f"Finalizing PDF: Writing binary to disk...")
                        log_container.code('\n'.join(st.session_state.logs), language='bash')
                        
                        clean_fn = f"{row['Download Filename ✏️'].split('.pdf')[0]}.pdf"
                        target_page.pdf(path=clean_fn, format="A4", print_background=True, margin={"top": "0.5in", "right": "0.5in", "bottom": "0.5in", "left": "0.5in"})
                        
                        update_excel_cell(row['Excel Row Index'], "Status", "✅ Downloaded")
                        add_log(f"💎 SUCCESS: '{clean_fn}' saved successfully.")
                        log_container.code('\n'.join(st.session_state.logs), language='bash')
                        print("--------------------------------------------------")
                    except Exception as e:
                        add_log(f"❌ ERROR (PDF): {str(e)[:80]}")
                        log_container.code('\n'.join(st.session_state.logs), language='bash')
                browser.close()
            add_log("=== HEAVYWEIGHT DOWNLOAD JOB COMPLETE ===")
            st.rerun()